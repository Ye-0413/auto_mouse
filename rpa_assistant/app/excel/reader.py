from __future__ import annotations

import logging
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

_logger = logging.getLogger(__name__)


@dataclass(slots=True)
class ExcelSheetSnapshot:
    """Subset of a worksheet for UI preview plus row counts."""

    sheet_names: list[str]
    headers: list[str]
    preview_rows: list[list[str]]
    """Trimmed to ``max_preview_rows`` for UI."""
    data_row_count: int
    """Number of non-empty data rows below the header row (full sheet scan)."""
    truncated_preview: bool


def _cell_str(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def open_workbook_meta(path: Path) -> list[str]:
    path = path.expanduser().resolve()
    if not path.is_file():
        raise FileNotFoundError(path)
    wb = load_workbook(filename=path, read_only=True, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def _non_empty_data_rows(ws: Worksheet, start_row: int) -> int:
    count = 0
    for row in ws.iter_rows(min_row=start_row, values_only=True):
        if any(cell not in (None, "") for cell in row):
            count += 1
    return count


def load_sheet_snapshot(
    path: Path,
    sheet_name: str,
    *,
    header_row_1based: int,
    max_preview_rows: int = 500,
) -> ExcelSheetSnapshot:
    """
    Load headers and up to ``max_preview_rows`` data rows, plus full data row count.

    ``header_row_1based`` is 1-based Excel row index for the header line.
    """
    path = path.expanduser().resolve()
    if not path.is_file():
        raise FileNotFoundError(path)
    if header_row_1based < 1:
        raise ValueError("header_row_1based must be >= 1")

    wb = load_workbook(filename=path, read_only=False, data_only=True)
    try:
        sheet_names = list(wb.sheetnames)
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet not found: {sheet_name!r}")
        ws = wb[sheet_name]

        header_cells = next(
            ws.iter_rows(
                min_row=header_row_1based,
                max_row=header_row_1based,
                values_only=True,
            ),
            (),
        )
        headers = [_cell_str(c) for c in header_cells]
        if not any(headers):
            _logger.warning("Header row %s appears empty", header_row_1based)

        start_data = header_row_1based + 1
        data_row_count = _non_empty_data_rows(ws, start_data)

        preview_rows: list[list[str]] = []
        truncated = False
        for row in ws.iter_rows(min_row=start_data, values_only=True):
            if len(preview_rows) >= max_preview_rows:
                truncated = True
                break
            values = list(row)[: len(headers)]
            if len(values) < len(headers):
                values.extend([None] * (len(headers) - len(values)))
            preview_rows.append([_cell_str(v) for v in values])

        return ExcelSheetSnapshot(
            sheet_names=sheet_names,
            headers=headers,
            preview_rows=preview_rows,
            data_row_count=data_row_count,
            truncated_preview=truncated,
        )
    finally:
        wb.close()
